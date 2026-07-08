"""Bangun laporan Excel (.xlsx) dari hasil prediksi TTP mapping.

Pemakaian:
    python src/build_excel_report.py [results.json] [output.xlsx]

- Jika results.json tidak diberikan, dipakai file terbaru di results/predictions/.
- Output default: results/metrics/laporan_ttp_<timestamp>.xlsx

Isi workbook:
    1. Ringkasan        : info run + metrik keseluruhan (teknik exact/base, taktik)
    2. Per Laporan      : GT vs prediksi teknik, TP/FP/FN, P/R/F1 per laporan
    3. Taktik per Lapor : taktik prediksi vs GT (diturunkan dari teknik), TP/FP/FN
    4. Distribusi Taktik: frekuensi tiap taktik diprediksi di seluruh laporan
"""
import os
import sys
import glob
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

# src/evaluation/build_excel_report.py -> tambahkan folder src ke path.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from knowledge.attck_loader import load_attck_techniques, load_attck_tactics
from evaluation.evaluator import (
    evaluate_predictions,
    evaluate_tactics,
    derive_tactic_ground_truth,
    _base_technique,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ATTCK_SOURCE = os.getenv("ATTCK_SOURCE", "data/mitre_cti/enterprise-attack.json")

# --- Gaya ---
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(bold=True, size=11, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")


def _latest_results() -> str:
    candidates = sorted(
        glob.glob("results/predictions/results_all_*.json")
        + glob.glob("results/predictions/eval_run_*.json")
        + glob.glob("results/predictions/results*.json"),
        key=os.path.getmtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("Tidak ada file hasil di results/predictions/")
    return candidates[0]


def _prf(tp: int, fp: int, fn: int) -> tuple[float, float, float]:
    precision = tp / (tp + fp) if (tp + fp) else 0.0
    recall = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) else 0.0
    return round(precision, 4), round(recall, 4), round(f1, 4)


def _style_header(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def _f1_fill(value: float) -> PatternFill | None:
    if value >= 0.5:
        return GOOD_FILL
    if value >= 0.2:
        return WARN_FILL
    if value > 0:
        return BAD_FILL
    return None


def build_summary_sheet(wb, results, attck_techniques, src_path):
    ws = wb.active
    ws.title = "Ringkasan"

    ws["A1"] = "Laporan Evaluasi TTP Mapping (MITRE ATT&CK)"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Sumber hasil"
    ws["B3"] = os.path.basename(src_path)
    ws["A4"] = "Knowledge base"
    ws["B4"] = os.path.basename(ATTCK_SOURCE)
    ws["A5"] = "Jumlah laporan"
    ws["B5"] = len(results)
    ws["A6"] = "Dibuat"
    ws["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for r in range(3, 7):
        ws.cell(row=r, column=1).font = Font(bold=True)

    tech = evaluate_predictions(results)
    tac = evaluate_tactics(results, attck_techniques)

    ws["A8"] = "Metrik Keseluruhan (micro-average)"
    ws["A8"].font = SUB_FONT

    headers = ["Kategori", "Precision", "Recall", "Micro-F1"]
    start = 9
    for j, h in enumerate(headers, 1):
        ws.cell(row=start, column=j, value=h)
    _style_header(ws, start, len(headers))

    rows = [
        ("Taktik (GT diturunkan dari teknik)", tac["tactic_precision"], tac["tactic_recall"], tac["tactic_micro_f1"]),
        ("Teknik — base (abaikan sub-teknik)", tech["base_precision"], tech["base_recall"], tech["base_micro_f1"]),
        ("Teknik — exact (termasuk sub-teknik)", tech["precision"], tech["recall"], tech["micro_f1"]),
    ]
    for i, (label, p, rec, f1) in enumerate(rows, start + 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=p)
        ws.cell(row=i, column=3, value=rec)
        cell_f1 = ws.cell(row=i, column=4, value=f1)
        fill = _f1_fill(f1)
        if fill:
            cell_f1.fill = fill
        for j in range(1, 5):
            ws.cell(row=i, column=j).border = BORDER
            if j > 1:
                ws.cell(row=i, column=j).alignment = CENTER

    ws["A14"] = "Keterangan warna F1: hijau ≥ 0.5, kuning ≥ 0.2, merah < 0.2"
    ws["A14"].font = Font(italic=True, size=9, color="808080")

    ws.column_dimensions["A"].width = 42
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14
    return tech, tac


def build_per_report_sheet(wb, results, attck_techniques):
    ws = wb.create_sheet("Per Laporan")
    headers = [
        "No", "Report ID", "#GT", "#Pred",
        "TP", "FP", "FN", "Precision", "Recall", "F1 (base)",
        "Teknik Prediksi", "Ground Truth (Teknik)",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for idx, r in enumerate(results, 1):
        gt = sorted({_base_technique(t) for t in r.get("ground_truth", [])})
        pred = sorted({_base_technique(t) for t in r.get("predicted_techniques", [])})
        gt_set, pred_set = set(gt), set(pred)
        tp = len(pred_set & gt_set)
        fp = len(pred_set - gt_set)
        fn = len(gt_set - pred_set)
        p, rec, f1 = _prf(tp, fp, fn)

        ws.append([
            idx,
            r.get("report_id", ""),
            len(gt_set), len(pred_set),
            tp, fp, fn, p, rec, f1,
            ", ".join(sorted(r.get("predicted_techniques", []))),
            ", ".join(sorted(r.get("ground_truth", []))),
        ])
        row = ws.max_row
        fill = _f1_fill(f1)
        if fill:
            ws.cell(row=row, column=10).fill = fill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            if col in (11, 12):
                cell.alignment = WRAP
            elif col >= 3:
                cell.alignment = CENTER

    widths = [5, 38, 6, 7, 5, 5, 5, 10, 9, 10, 45, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_tactic_sheet(wb, results, attck_techniques):
    ws = wb.create_sheet("Taktik per Laporan")
    tac_names = load_attck_tactics(ATTCK_SOURCE)
    headers = [
        "No", "Report ID", "TP", "FP", "FN", "Precision", "Recall", "F1",
        "Taktik Prediksi", "Taktik Ground Truth (diturunkan)",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    def _fmt(ids):
        return ", ".join(f"{t} ({tac_names.get(t, '?')})" for t in ids)

    for idx, r in enumerate(results, 1):
        gt = set(derive_tactic_ground_truth(r.get("ground_truth", []), attck_techniques))
        pred = set(r.get("tactics_identified", []))
        tp = len(pred & gt)
        fp = len(pred - gt)
        fn = len(gt - pred)
        p, rec, f1 = _prf(tp, fp, fn)
        ws.append([
            idx, r.get("report_id", ""), tp, fp, fn, p, rec, f1,
            _fmt(sorted(pred)), _fmt(sorted(gt)),
        ])
        row = ws.max_row
        fill = _f1_fill(f1)
        if fill:
            ws.cell(row=row, column=8).fill = fill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            if col in (9, 10):
                cell.alignment = WRAP
            elif col >= 3:
                cell.alignment = CENTER

    widths = [5, 38, 5, 5, 5, 10, 9, 9, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_distribution_sheet(wb, results):
    ws = wb.create_sheet("Distribusi Taktik")
    tac_names = load_attck_tactics(ATTCK_SOURCE)
    counter = Counter()
    for r in results:
        for t in r.get("tactics_identified", []):
            counter[t] += 1

    headers = ["Tactic ID", "Nama", "Jumlah Laporan", "% dari total"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    total = len(results) or 1
    for tid, name in tac_names.items():
        cnt = counter.get(tid, 0)
        ws.append([tid, name, cnt, round(100 * cnt / total, 1)])
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = BORDER
            if col >= 3:
                ws.cell(row=row, column=col).alignment = CENTER

    widths = [12, 26, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main():
    args = [a for a in sys.argv[1:]]
    results_path = args[0] if len(args) >= 1 else _latest_results()
    if len(args) >= 2:
        out_path = args[1]
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs("results/metrics", exist_ok=True)
        out_path = f"results/metrics/laporan_ttp_{ts}.xlsx"

    with open(results_path, "r", encoding="utf-8") as f:
        results = json.load(f)
    print(f"Memuat {len(results)} hasil dari {results_path}")

    attck_techniques = load_attck_techniques(ATTCK_SOURCE)

    wb = Workbook()
    tech, tac = build_summary_sheet(wb, results, attck_techniques, results_path)
    build_per_report_sheet(wb, results, attck_techniques)
    build_tactic_sheet(wb, results, attck_techniques)
    build_distribution_sheet(wb, results)

    wb.save(out_path)
    print(f"Laporan Excel disimpan ke: {out_path}")
    print(f"  Taktik  F1={tac['tactic_micro_f1']}  | Teknik base F1={tech['base_micro_f1']}  exact F1={tech['micro_f1']}")


if __name__ == "__main__":
    main()
